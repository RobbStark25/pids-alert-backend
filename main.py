from fastapi import FastAPI, Request, Header, Depends, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from pydantic import BaseModel
from collections import Counter
from datetime import datetime, timedelta
import sqlite3
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import requests
import json
import os
import io
import threading
import time
from fastapi import Body
import subprocess
import platform
import openpyxl
from fastapi import Depends
import csv
from io import BytesIO
from fastapi import APIRouter
from typing import List
from datetime import datetime, timedelta

LINEWALKER_FILE = "linewalkers.json"
RESET_DURATION_HOURS = 9

app = FastAPI()
sent_count = 0

API_KEY = "Yj@mb51"
DB_FILE = "log.sqlite"
SETTINGS_FILE = "settings.json"
EXCEL_EXPORT_FILE = "PIDS_Log_Export.xlsx"  # or a full path
LINEWALKER_FILE = "linewalkers.json"

settings = {}
received_messages = []
last_update_id = None

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def verify_api_key(x_api_key: str = Header(...)):
    if x_api_key != API_KEY:
        raise HTTPException(status_code=403, detail="Unauthorized")

def load_settings():
    global settings
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE) as f:
            settings = json.load(f)
    else:
        settings = {"BOT_TOKEN": "", "CHAT_ID": ""}
load_settings()


class TokenData(BaseModel):
    token: str
    chat_id: str

@app.post("/update_token")
def update_token(data: TokenData, auth=Depends(verify_api_key)):
    try:
        config = {"BOT_TOKEN": data.token, "CHAT_ID": data.chat_id}
        with open(SETTINGS_FILE, "w") as f:
            json.dump(config, f, indent=2)
        global settings
        settings = load_settings()
        return {"status": "success", "message": "Token updated"}
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "detail": str(e)})


def log_message_sqlite(data):
    now = datetime.now()
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''INSERT INTO sent_logs (date, time, od, ch, section, linewalker)
                 VALUES (?, ?, ?, ?, ?, ?)''', (
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        data.get("OD", ""),
        data.get("CH", ""),
        data.get("Section", ""),
        data.get("LineWalker", "")
    ))
    conn.commit()
    conn.close()

# ========== Settings and Linewalkers ==========

def load_linewalkers():
    if os.path.exists(LINEWALKER_FILE):
        with open(LINEWALKER_FILE) as f:
            return json.load(f)
    return []

def save_linewalkers(data):
    with open(LINEWALKER_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def refresh_linewalkers():
    global linewalker_data
    linewalker_data = load_linewalkers()

linewalker_data = load_linewalkers()

# ========== Section Data ==========
section_files = {
    "IPS to SV-08": "OD_CH_1.csv",
    "SV-09 to SV-08": "OD_CH_2.csv",
    "SV-09 to SV-10": "OD_CH_3.csv",
    "SV-11 to SV-10": "OD_CH_4.csv",
    "SV-11 to KRS": "OD_CH_5.csv"
}
section_data = {}
for section, file in section_files.items():
    try:
        df = pd.read_csv(file)
        df = df.dropna(subset=["OD", "CH"])
        df = df.sort_values("OD")
        df["Diff"] = df["OD"].diff().fillna(0)
        section_data[section] = df.reset_index(drop=True)
    except Exception as e:
        print(f"Error loading {file} for section {section}: {e}")

# ========== Interpolation ==========
def interpolate_ch(df, od):
    ch_matches = []
    for i in range(len(df) - 1):
        od1 = df.loc[i, "OD"]
        od2 = df.loc[i + 1, "OD"]
        ch1 = df.loc[i, "CH"]
        ch2 = df.loc[i + 1, "CH"]
        diff = df.loc[i, "Diff"]
        if od1 <= od <= od2 and diff != 0:
            od_diff = od - od1
            ch = ch1 + ((ch2 - ch1) * od_diff / diff)
            ch_matches.append(round(ch, 3))
    return ch_matches  # Always returns a list

def interpolate_od(df, ch):
    for i in range(len(df) - 1):
        ch1 = df.loc[i, "CH"]
        ch2 = df.loc[i + 1, "CH"]
        od1 = df.loc[i, "OD"]
        od2 = df.loc[i + 1, "OD"]
        if ch1 <= ch <= ch2:
            interpolated = od1 + ((ch - ch1) * (od2 - od1)) / (ch2 - ch1)
            return round(interpolated)
    return None

def get_linewalker_by_ch(ch):
    for entry in linewalker_data:
        if entry["start_ch"] <= ch <= entry["end_ch"]:
            return entry["line_walker"]
    return None

# ========== Main API ==========
@app.get("/calculate_ch_for_section")
def calculate_ch_for_section(section: str, od: float):
    print(f"[CH Lookup] Section={section}, OD={od}")

    df = section_data.get(section)
    if df is None:
        return {"error": f"Section '{section}' not found."}

    ch_matches = interpolate_ch(df, od)

    if not ch_matches:
        return {"error": "OD out of range or no valid interpolation found."}

    # ✅ Multiple CHs — return list directly
    if len(ch_matches) > 1:
        print(f"[Multiple CHs] Found: {ch_matches}")
        return ch_matches

    # ✅ Single CH — return with linewalker
    ch_val = ch_matches[0]
    lw = get_linewalker_by_ch(ch_val)
    if not lw:
        return {"error": "Line walker not found for CH."}
    
    return {
        "ch": ch_val,
        "line_walker": lw
    }



@app.get("/convert/ch-to-od")
def convert_ch_to_od(section: str, ch: float):
    df = section_data.get(section)
    if df is None:
        return {"error": f"Section '{section}' not found."}
    od = interpolate_od(df, ch)
    if od is None:
        return {"error": "CH out of range."}
    return {"od": od}

class AlertPayload(BaseModel):
    od: float
    ch: float
    section: str
    line_walker: str

@app.post("/send_alert")
def send_alert(payload: AlertPayload):
    msg = (
        "🔔 अलार्म सूचना 🔔\n"
        f"⏱️समय: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}\n"
        f"🔎OD: {payload.od}\n"
        f"📍CH: {payload.ch}\n"
        f"📈सेक्शन: {payload.section}\n"
        f"🚶‍➡️लाइन वॉकर: {payload.line_walker}"
    )

    url = f"https://api.telegram.org/bot{settings['BOT_TOKEN']}/sendMessage"

    try:
        res = requests.post(url, json={"chat_id": settings['CHAT_ID'], "text": msg})
        if res.status_code == 200:
            # ✅ Log to SQLite
            log_message_sqlite({
                "OD": payload.od,
                "CH": payload.ch,
                "Section": payload.section,
                "LineWalker": payload.line_walker
            })
            return {
                "status": "success",
                "message_id": res.json().get("result", {}).get("message_id", None)
            }
        else:
            return {
                "status": "error",
                "detail": f"Telegram API returned {res.status_code}: {res.text}"
            }

    except Exception as e:
        return {
            "status": "error",
            "detail": f"Exception while sending alert: {str(e)}"
        }

@app.get("/receive")
def get_received_logs(limit: int = 100):
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT timestamp, linewalker, message, user 
            FROM received_messages 
            ORDER BY timestamp DESC 
            LIMIT ?
        """, (limit,))
        rows = c.fetchall()
        conn.close()

        return [
            {
                "User": row[3],  # user
                "Message": row[2],  # message
                "Time": row[0].split(" ")[1] if " " in row[0] else row[0]  # extract HH:MM:SS
            }
            for row in rows
        ]

    except Exception as e:
        print(f"[DB Error] {e}")
        return {"error": str(e)}

# ========== Logging ==========
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # Table: Sent Logs
    c.execute('''
        CREATE TABLE IF NOT EXISTS sent_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            time TEXT,
            od REAL,
            ch TEXT,
            section TEXT,
            linewalker TEXT
        )
    ''')

    # Table: Received Messages
    c.execute('''
        CREATE TABLE IF NOT EXISTS received_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            linewalker TEXT,
            message TEXT,
            user TEXT
        )
    ''')

    # Table: Duty Status with separate columns for ON and OFF messages
    c.execute('''
        CREATE TABLE IF NOT EXISTS duty_status (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            linewalker TEXT,
            duty_on TEXT,
            duty_off TEXT
        )
    ''')

    conn.commit()
    conn.close()

init_db()


@app.post("/webhook")
async def webhook(request: Request, background_tasks: BackgroundTasks):
    data = await request.json()
    background_tasks.add_task(handle_webhook, data)
    return {"status": "received"}

def handle_webhook(data):
    try:
        msg = data.get("message", {})
        text = msg.get("text", "").strip()
        user = msg.get("from", {}).get("first_name", "Unknown")

        if text:
            log_duty_status_from_message(user, text, user)

    except Exception as e:
        print("Webhook error:", e)



def log_duty_status_from_message(linewalker, message, user):
    msg_lower = message.lower()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    duty_on_msg = message if "on" in msg_lower and "off" not in msg_lower else None
    duty_off_msg = message if "off" in msg_lower else None

    if duty_on_msg or duty_off_msg:
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute('''
                INSERT INTO duty_status (timestamp, linewalker, duty_on, duty_off)
                VALUES (?, ?, ?, ?)
            ''', (timestamp, linewalker, duty_on_msg, duty_off_msg))
            conn.commit()
            print(f"[LOG] Duty status logged for {linewalker}")
        except Exception as e:
            print(f"[Log Error] Failed to log duty status: {e}")
        finally:
            conn.close()
    else:
        log_received_message(linewalker, message, user)


def log_received_message(linewalker, message, user):
    now = datetime.now()
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('''
            INSERT INTO received_messages (timestamp, linewalker, message, user)
            VALUES (?, ?, ?, ?)
        ''', (
            now.strftime("%Y-%m-%d %H:%M:%S"),
            linewalker,
            message,
            user
        ))
        conn.commit()
        print(f"[✓] Logged general message from {linewalker}")
    except Exception as e:
        print(f"[Log Error] Failed to log received message: {e}")
    finally:
        conn.close()


def clear_duty_status_if_due():
    last_cleared_date = None
    while True:
        now = datetime.now()
        if now.strftime("%H:%M") == "06:30" and last_cleared_date != now.strftime("%Y-%m-%d"):
            try:
                conn = sqlite3.connect(DB_FILE)
                c = conn.cursor()
                c.execute("DELETE FROM duty_status")
                conn.commit()
                conn.close()
                print(f"[✓] Duty_Status auto-cleared at 06:30 on {now.strftime('%Y-%m-%d')}")
                last_cleared_date = now.strftime("%Y-%m-%d")
            except Exception as e:
                print(f"[!] Error clearing Duty_Status: {e}")
        time.sleep(60)


threading.Thread(target=clear_duty_status_if_due, daemon=True).start()

@app.get("/set_webhook")
def set_webhook():
    try:
        url = f"https://api.telegram.org/bot{settings['BOT_TOKEN']}/setWebhook"
        webhook_url = "https://pids-alert-backend.onrender.com/webhook"
        res = requests.get(url, params={"url": webhook_url})
        return res.json()
    except Exception as e:
        return {"status": "error", "detail": str(e)}

# ============== View Logs and Export to Excel ============
@app.get("/view_logs")
def view_logs():
    try:
        conn = sqlite3.connect(DB_FILE)
        df_recv = pd.read_sql("SELECT * FROM received_messages", conn)
        df_duty = pd.read_sql("SELECT * FROM duty_status", conn)
        df_sent = pd.read_sql("SELECT * FROM sent_logs", conn)
        conn.close()

        # Create Excel file
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Received Logs"
        ws1.append(df_recv.columns.tolist())
        for row in df_recv.itertuples(index=False):
            ws1.append(list(row))

        ws2 = wb.create_sheet(title="Duty Status")
        ws2.append(df_duty.columns.tolist())
        for row in df_duty.itertuples(index=False):
            ws2.append(list(row))

        ws3 = wb.create_sheet(title="Sent Logs")
        ws3.append(df_sent.columns.tolist())
        for row in df_sent.itertuples(index=False):
            ws3.append(list(row))

        wb.save(EXCEL_EXPORT_FILE)
        os.startfile(EXCEL_EXPORT_FILE)  # or subprocess on Mac/Linux
        return {"status": "opened_excel", "file": EXCEL_EXPORT_FILE}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/download_logs")
def download_logs():
    try:
        # Reuse same code to generate Excel
        conn = sqlite3.connect(DB_FILE)
        df_recv = pd.read_sql("SELECT * FROM received_messages", conn)
        df_duty = pd.read_sql("SELECT * FROM duty_status", conn)
        df_sent = pd.read_sql("SELECT * FROM sent_logs", conn)
        conn.close()

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Received Logs"
        ws1.append(df_recv.columns.tolist())
        for row in df_recv.itertuples(index=False):
            ws1.append(list(row))

        ws2 = wb.create_sheet(title="Duty Status")
        ws2.append(df_duty.columns.tolist())
        for row in df_duty.itertuples(index=False):
            ws2.append(list(row))

        ws3 = wb.create_sheet(title="Sent Logs")
        ws3.append(df_sent.columns.tolist())
        for row in df_sent.itertuples(index=False):
            ws3.append(list(row))

        wb.save(EXCEL_EXPORT_FILE)

        return FileResponse(
            EXCEL_EXPORT_FILE,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename="PIDS_Log_Export.xlsx"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ================ Analytics Charts from SQLite ===============
@app.get("/analytics/scatter_chart")
def get_scatter_chart():
    try:
        conn = sqlite3.connect(DB_FILE)
        df = pd.read_sql("SELECT * FROM sent_logs", conn)
        conn.close()

        df['datetime'] = pd.to_datetime(df['date'] + ' ' + df['time'])
        now = datetime.now()
        today_630 = now.replace(hour=6, minute=30, second=0, microsecond=0)
        if now < today_630:
            today_630 -= timedelta(days=1)
        tomorrow_630 = today_630 + timedelta(days=1)
        df = df[(df['datetime'] >= today_630) & (df['datetime'] < tomorrow_630)]

        fig, ax = plt.subplots(figsize=(11, 6))
        for section, group in df.groupby("section"):
            ax.scatter(group['datetime'], group['ch'].astype(float), label=section, s=40, alpha=0.8)

        ax.set_title("📊 Chainage vs Time (Section-wise, Last 24 Hours)", fontsize=14)
        ax.set_xlabel("Time", fontsize=12)
        ax.set_ylabel("Chainage", fontsize=12)
        ax.legend(title="Section", loc="best")
        ax.grid(True)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        fig.autofmt_xdate()

        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        plt.close(fig)
        buf.seek(0)
        return StreamingResponse(buf, media_type="image/png")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/analytics/grouping_chart")
def get_grouping_chart(by: str = "section"):
    try:
        valid_fields = {
            "section": "section",
            "linewalker": "linewalker",
    
        }
        if by not in valid_fields:
            raise HTTPException(status_code=400, detail=f"Invalid group type. Use one of: {', '.join(valid_fields.keys())}")

        conn = sqlite3.connect(DB_FILE)
        df = pd.read_sql("SELECT * FROM sent_logs", conn)
        conn.close()

        df['datetime'] = pd.to_datetime(df['date'] + ' ' + df['time'])
        now = datetime.now()
        today_630 = now.replace(hour=6, minute=30, second=0, microsecond=0)
        if now < today_630:
            today_630 -= timedelta(days=1)
        tomorrow_630 = today_630 + timedelta(days=1)
        df = df[(df['datetime'] >= today_630) & (df['datetime'] < tomorrow_630)]

        counts = df[by].value_counts()

        fig, ax = plt.subplots(figsize=(10, 6))
        ax.bar(counts.index.astype(str), counts.values, color='teal')
        ax.set_title(f"Alarm Count by {by.capitalize()} (Last 24 Hours)", fontsize=14)
        ax.set_ylabel("Count")
        ax.set_xlabel(by.capitalize())
        ax.set_xticks(range(len(counts)))
        ax.set_xticklabels(counts.index.astype(str), rotation=45, ha='right')
        ax.grid(axis="y")

        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format="png")
        plt.close(fig)
        buf.seek(0)
        return StreamingResponse(buf, media_type="image/png")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ========== Linewalker Management ==========


class LineWalkerItem(BaseModel):
    start_ch: float
    end_ch: float
    line_walker: str

# ✅ View all linewalkers
@app.get("/view_linewalkers")
def view_linewalkers():
    return load_linewalkers()

# ✅ Update linewalkers via frontend
@app.post("/edit_linewalkers")
def edit_linewalkers(data: list[LineWalkerItem]):
    try:
        # Convert to dict list
        data_dicts = [item.dict() for item in data]
        save_linewalkers(data_dicts)

        # 🔁 Update in-memory variable if used
        global linewalker_data
        linewalker_data = data_dicts

        return {"status": "updated"}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

# ✅ Save linewalkers to file
def save_linewalkers(data):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for item in data:
        item["saved_at"] = now_str

    with open(LINEWALKER_FILE, 'w') as f:
        json.dump(data, f, indent=2)

# ✅ Load and auto-expire linewalkers
def load_linewalkers():
    if not os.path.exists(LINEWALKER_FILE):
        return []

    with open(LINEWALKER_FILE) as f:
        data = json.load(f)

    now = datetime.now()
    updated = False

    for item in data:
        saved_time_str = item.get("saved_at")
        if saved_time_str:
            try:
                saved_time = datetime.strptime(saved_time_str, "%Y-%m-%d %H:%M:%S")
                if now - saved_time > timedelta(hours=RESET_DURATION_HOURS):
                    item["line_walker"] = "-"
                    item["saved_at"] = None
                    updated = True
            except Exception as e:
                print(f"Invalid saved_at format: {saved_time_str}. Skipping reset.")

    if updated:
        with open(LINEWALKER_FILE, "w") as f:
            json.dump(data, f, indent=2)

    return data

# Optional refresh endpoint
@app.get("/refresh_linewalkers")
def refresh_linewalkers_api():
    refreshed = load_linewalkers()
    return {"status": "refreshed", "count": len(refreshed)}

@app.post("/reset_all_linewalkers")
def reset_all_linewalkers():
    if not os.path.exists(LINEWALKER_FILE):
        return {"status": "file_not_found"}

    with open(LINEWALKER_FILE, "r") as f:
        data = json.load(f)

    for item in data:
        item["line_walker"] = "-"
        item["saved_at"] = None

    with open(LINEWALKER_FILE, "w") as f:
        json.dump(data, f, indent=2)

    # update global variable if used
    global linewalker_data
    linewalker_data = data

    return {"status": "reset", "count": len(data)}


@app.get("/ping")
def ping():
    return {"status": "ok"}
@app.get("/")
def root():
    return {"message": "✅ PIDS Alert Backend is Running"}

